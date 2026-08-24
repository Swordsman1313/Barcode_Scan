"""
Comprehensive Automated Test Suite for Telegram Stock Count Bot.
Tests:
1. Database initialization and high-concurrency multi-user inserts (Simulating 30 concurrent crew members).
2. Sticky shelf preferences per user.
3. Summary stats aggregation.
4. Excel .xlsx file generation, styling, and text-safe barcode validation.
5. Barcode auto-detection with image synthesis and zxingcpp.
6. Google Sheet async sync worker queue.
"""

import asyncio
import os
import sys
import unittest
from pathlib import Path
from PIL import Image
import openpyxl
import zxingcpp

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import database
import config
from excel_exporter import create_excel_report
from barcode_reader import detect_barcode_from_image
from sheets_sync import SheetsSyncManager


class TestStockBotSystem(unittest.IsolatedAsyncioTestCase):

    async def asyncSetUp(self):
        # Use a temporary test database
        config.DATABASE_PATH = "test_inventory.db"
        if os.path.exists("test_inventory.db"):
            os.remove("test_inventory.db")
        await database.init_db()

    async def asyncTearDown(self):
        if os.path.exists("test_inventory.db"):
            os.remove("test_inventory.db")

    async def test_concurrent_crew_inserts(self):
        """Simulates 30 store crew members simultaneously logging stock items."""
        async def mock_crew_worker(user_id: int, crew_name: str, shelf: str, count_num: int):
            for i in range(count_num):
                await database.insert_count(
                    user_id=user_id,
                    crew_name=crew_name,
                    shelf=shelf,
                    barcode=f"885{user_id:04d}{i:05d}",
                    item_name=f"Product SKU {i} by {crew_name}",
                    qty=(i + 1) * 2
                )

        # 30 concurrent crew tasks
        tasks = []
        for uid in range(1, 31):
            shelf = f"G10{uid % 5 + 1}"
            tasks.append(mock_crew_worker(uid, f"Crew_Member_{uid}", shelf, 5))

        await asyncio.gather(*tasks)

        # Verify total inserted counts = 30 * 5 = 150 items
        all_counts = await database.get_all_counts()
        self.assertEqual(len(all_counts), 150, "All 150 concurrent records should be saved without data loss.")

        # Test stats
        stats = await database.get_summary_stats()
        self.assertEqual(stats["total_skus"], 150)
        self.assertGreater(stats["total_qty"], 0)
        self.assertEqual(len(stats["crew_breakdown"]), 30)

    async def test_sticky_shelf_preference(self):
        """Tests that active shelf is remembered per crew member."""
        await database.set_user_active_shelf(1001, "A-12")
        await database.set_user_active_shelf(1002, "B-05")

        shelf_1 = await database.get_user_active_shelf(1001)
        shelf_2 = await database.get_user_active_shelf(1002)

        self.assertEqual(shelf_1, "A-12")
        self.assertEqual(shelf_2, "B-05")

    async def test_excel_export_generation(self):
        """Tests Excel generation with multi-tab workbook, styled columns, and text-safe barcodes."""
        sample_counts = [
            {
                "id": 1,
                "timestamp": "2026-08-24 10:00:00",
                "crew_name": "Somchai (Crew A)",
                "shelf": "G101",
                "barcode": "08850123456789",  # Note leading zero and 14 digits
                "item_name": "Coca Cola 325ml",
                "qty": 24,
                "synced_sheet": 1
            },
            {
                "id": 2,
                "timestamp": "2026-08-24 10:05:00",
                "crew_name": "Nok (Crew B)",
                "shelf": "G102",
                "barcode": "8851234567890",
                "item_name": "Lay's Nori Seaweed 50g",
                "qty": 12,
                "synced_sheet": 0
            }
        ]

        excel_buf = create_excel_report(sample_counts)
        self.assertIsNotNone(excel_buf)
        self.assertGreater(excel_buf.getbuffer().nbytes, 1000)

        # Load workbook and verify sheets & cell formats
        wb = openpyxl.load_workbook(excel_buf)
        self.assertIn("Stock Items", wb.sheetnames)
        self.assertIn("Summary by Shelf", wb.sheetnames)
        self.assertIn("Summary by Crew", wb.sheetnames)

        ws_items = wb["Stock Items"]
        # Row 3 is first data row
        barcode_cell = ws_items.cell(row=3, column=5)
        self.assertEqual(barcode_cell.value, "08850123456789")
        self.assertEqual(barcode_cell.number_format, "@", "Barcode must be formatted as text string (@)")

    def test_barcode_auto_reader(self):
        """Generates a test EAN13 barcode image and tests detection."""
        barcode_str = "8850123456787"
        barcode_obj = zxingcpp.create_barcode(barcode_str, zxingcpp.BarcodeFormat.EAN13)
        zx_img = zxingcpp.write_barcode_to_image(barcode_obj)

        pil_img = Image.frombuffer("L", (zx_img.shape[1], zx_img.shape[0]), zx_img)
        test_img_path = "test_barcode_sample.png"
        pil_img.save(test_img_path)

        try:
            detected = detect_barcode_from_image(test_img_path)
            self.assertEqual(detected, barcode_str, "Barcode reader should correctly detect barcode from image")
        finally:
            if os.path.exists(test_img_path):
                os.remove(test_img_path)

    async def test_sync_manager_queue(self):
        """Tests that async Sync Manager enqueues items without raising errors."""
        manager = SheetsSyncManager(webhook_url=None)  # None URL mode
        await manager.start()
        
        manager.enqueue({
            "id": 1,
            "timestamp": "2026-08-24 10:00:00",
            "crew_name": "Test User",
            "shelf": "G101",
            "barcode": "885000000000",
            "item_name": "Test Item",
            "qty": 5
        })

        await asyncio.sleep(0.1)
        await manager.stop()


if __name__ == "__main__":
    unittest.main()

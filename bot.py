"""
=============================================================================
STORE STOCK COUNT TELEGRAM BOT — AI VISION PACKAGING NAME DETECTOR
=============================================================================
- Auto-catches Product Name directly from packaging photo (Gemini AI Vision)
- Auto-detects Barcodes from photos (zxing-cpp)
- Manual Shelf & Quantity input
- Smart "Keep Last Shelf" button for fast repeat scans
- Real-time Google Sheets Auto-Fill via Webhook
- Instant Excel (.xlsx) Export
- 24/7 Cloud Ready
=============================================================================
"""

import os
import re
import uuid
import base64
import asyncio
import logging
import zoneinfo
from io import BytesIO
from datetime import datetime
from pathlib import Path
from contextlib import asynccontextmanager
from typing import Optional, List, Dict, Any, Tuple

from dotenv import load_dotenv
from PIL import Image, ImageEnhance, ImageOps

try:
    import zxingcpp
    HAS_ZXING = True
except ImportError:
    HAS_ZXING = False

import aiosqlite
import aiohttp
from aiohttp import web
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    KeyboardButton
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters
)

# ---------------------------------------------------------------------------
# 1. CONFIGURATION
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
GOOGLE_SHEET_WEBHOOK_URL = os.getenv("GOOGLE_SHEET_WEBHOOK_URL", "").strip()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
DATABASE_PATH = os.getenv("DATABASE_PATH", str(BASE_DIR / "inventory.db"))
PHOTOS_DIR = Path(os.getenv("PHOTOS_DIR", str(BASE_DIR / "photos")))
PHOTOS_DIR.mkdir(parents=True, exist_ok=True)
TIMEZONE = os.getenv("TIMEZONE", "Asia/Bangkok")
PORT = int(os.getenv("PORT", "8080"))

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger("StockBot")

(
    STATE_BARCODE_PHOTO,
    STATE_SHELF,
    STATE_BARCODE,
    STATE_ITEM_NAME,
    STATE_QTY
) = range(5)


# ---------------------------------------------------------------------------
# 2. KEYBOARDS
# ---------------------------------------------------------------------------
def get_main_reply_keyboard() -> ReplyKeyboardMarkup:
    keyboard = [
        [KeyboardButton("📸 Count New Item"), KeyboardButton("📍 Set Shelf")],
        [KeyboardButton("📊 Export Excel"), KeyboardButton("📈 View Stats")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


# ---------------------------------------------------------------------------
# 3. DATABASE LAYER (SQLite WAL Mode)
# ---------------------------------------------------------------------------
def get_current_timestamp() -> str:
    try:
        tz = zoneinfo.ZoneInfo(TIMEZONE)
        return datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@asynccontextmanager
async def get_db():
    db = await aiosqlite.connect(DATABASE_PATH)
    db.row_factory = aiosqlite.Row
    await db.execute("PRAGMA journal_mode=WAL;")
    await db.execute("PRAGMA synchronous=NORMAL;")
    await db.execute("PRAGMA busy_timeout=10000;")
    try:
        yield db
    finally:
        await db.close()


async def init_db():
    async with get_db() as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS counts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                crew_name TEXT NOT NULL,
                shelf TEXT NOT NULL,
                barcode TEXT NOT NULL,
                item_name TEXT,
                qty REAL NOT NULL DEFAULT 1,
                photo_front TEXT,
                photo_barcode TEXT,
                synced_sheet INTEGER NOT NULL DEFAULT 0
            );
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS user_preferences (
                user_id INTEGER PRIMARY KEY,
                active_shelf TEXT,
                updated_at TEXT
            );
        """)
        await db.execute("CREATE INDEX IF NOT EXISTS idx_counts_shelf ON counts(shelf);")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_counts_user ON counts(user_id);")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_counts_barcode ON counts(barcode);")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_counts_synced ON counts(synced_sheet);")
        await db.commit()


async def db_insert_count(user_id: int, crew_name: str, shelf: str, barcode: str, item_name: str, qty: float, photo_front: str = None, photo_barcode: str = None) -> Dict[str, Any]:
    timestamp = get_current_timestamp()
    shelf_clean = (shelf or "UNKNOWN").strip().upper()
    barcode_clean = str(barcode or "NO_BARCODE").strip()
    item_name_clean = (item_name or "-").strip()

    async with get_db() as db:
        cursor = await db.execute(
            """
            INSERT INTO counts (
                timestamp, user_id, crew_name, shelf, barcode, item_name, qty, photo_front, photo_barcode, synced_sheet
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
            """,
            (timestamp, user_id, crew_name, shelf_clean, barcode_clean, item_name_clean, qty, photo_front, photo_barcode)
        )
        row_id = cursor.lastrowid
        await db.commit()

    return {
        "id": row_id,
        "timestamp": timestamp,
        "user_id": user_id,
        "crew_name": crew_name,
        "shelf": shelf_clean,
        "barcode": barcode_clean,
        "item_name": item_name_clean,
        "qty": qty,
        "photo_front": photo_front,
        "photo_barcode": photo_barcode,
        "synced_sheet": 0
    }


async def db_get_user_active_shelf(user_id: int) -> Optional[str]:
    async with get_db() as db:
        async with db.execute("SELECT active_shelf FROM user_preferences WHERE user_id = ?", (user_id,)) as cursor:
            row = await cursor.fetchone()
            if row and row["active_shelf"]:
                return row["active_shelf"]
    return None


async def db_set_user_active_shelf(user_id: int, shelf: str):
    timestamp = get_current_timestamp()
    shelf_clean = shelf.strip().upper()
    async with get_db() as db:
        await db.execute(
            """
            INSERT INTO user_preferences (user_id, active_shelf, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
                active_shelf = excluded.active_shelf,
                updated_at = excluded.updated_at
            """,
            (user_id, shelf_clean, timestamp)
        )
        await db.commit()


async def db_get_all_counts() -> List[Dict[str, Any]]:
    async with get_db() as db:
        async with db.execute("SELECT * FROM counts ORDER BY id ASC") as cursor:
            rows = await cursor.fetchall()
            return [dict(r) for r in rows]


async def db_mark_synced(count_ids: List[int]):
    if not count_ids:
        return
    placeholders = ",".join("?" for _ in count_ids)
    async with get_db() as db:
        await db.execute(f"UPDATE counts SET synced_sheet = 1 WHERE id IN ({placeholders})", count_ids)
        await db.commit()


async def db_get_unsynced_counts(limit: int = 50) -> List[Dict[str, Any]]:
    async with get_db() as db:
        async with db.execute("SELECT * FROM counts WHERE synced_sheet = 0 ORDER BY id ASC LIMIT ?", (limit,)) as cursor:
            rows = await cursor.fetchall()
            return [dict(r) for r in rows]


async def db_get_summary_stats() -> Dict[str, Any]:
    async with get_db() as db:
        async with db.execute("SELECT COUNT(*) as total_skus, COALESCE(SUM(qty), 0) as total_qty, COUNT(DISTINCT shelf) as total_shelves FROM counts") as c1:
            r1 = await c1.fetchone()
            total_skus = r1["total_skus"] if r1 else 0
            total_qty = r1["total_qty"] if r1 else 0
            total_shelves = r1["total_shelves"] if r1 else 0

        async with db.execute("SELECT shelf, COUNT(*) as sku_count, SUM(qty) as total_qty FROM counts GROUP BY shelf ORDER BY shelf ASC") as c2:
            shelf_breakdown = [dict(r) for r in await c2.fetchall()]

        async with db.execute("SELECT crew_name, COUNT(*) as sku_count, SUM(qty) as total_qty FROM counts GROUP BY user_id, crew_name ORDER BY total_qty DESC") as c3:
            crew_breakdown = [dict(r) for r in await c3.fetchall()]

        async with db.execute("SELECT COUNT(*) as pending FROM counts WHERE synced_sheet = 0") as c4:
            r4 = await c4.fetchone()
            pending_sync = r4["pending"] if r4 else 0

    return {
        "total_skus": total_skus,
        "total_qty": total_qty,
        "total_shelves": total_shelves,
        "pending_sync": pending_sync,
        "shelf_breakdown": shelf_breakdown,
        "crew_breakdown": crew_breakdown
    }


# ---------------------------------------------------------------------------
# 4. AI VISION PRODUCT NAME EXTRACTOR (Auto-Catch Name from Packaging)
# ---------------------------------------------------------------------------
async def extract_product_name_from_image(image_path: str) -> Optional[str]:
    """
    Uses Gemini AI Vision to automatically detect product name from packaging.
    Returns concise Brand & Product Name (e.g. 'Jardo Seaweed Rice Chip', 'Coca Cola 325ml').
    """
    if not image_path or not os.path.exists(image_path):
        return None

    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        return None

    try:
        with open(image_path, "rb") as f:
            img_b64 = base64.b64encode(f.read()).decode("utf-8")

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        payload = {
            "contents": [{
                "parts": [
                    {
                        "text": (
                            "Look at this product packaging image. Identify and extract ONLY the Brand and Product Name "
                            "(including flavor or size if visible). Return ONLY the product name (maximum 6 words). "
                            "Do not include markdown, quotes, bullet points, or filler words. "
                            "Example output: 'Jardo Seaweed Rice Chip' or 'Lay's Classic 50g' or 'Coca Cola 325ml'."
                        )
                    },
                    {
                        "inlineData": {
                            "mimeType": "image/jpeg",
                            "data": img_b64
                        }
                    }
                ]
            }]
        }

        async with aiohttp.ClientSession() as session:
            async with session.post(url, json=payload, timeout=aiohttp.ClientTimeout(total=8)) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    candidates = data.get("candidates", [])
                    if candidates:
                        text = candidates[0].get("content", {}).get("parts", [{}])[0].get("text", "")
                        clean_name = text.strip().replace("\n", " ").replace("*", "").replace('"', '').strip()
                        if clean_name and len(clean_name) > 2:
                            logger.info(f"✨ AI detected product name: {clean_name}")
                            return clean_name
                else:
                    err_text = await resp.text()
                    logger.warning(f"Gemini API returned status {resp.status}: {err_text}")
    except Exception as e:
        logger.warning(f"Error extracting product name with AI: {e}")

    return None


# ---------------------------------------------------------------------------
# 5. BARCODE RECOGNITION (zxing-cpp + PIL)
# ---------------------------------------------------------------------------
def detect_barcode_from_image(image_path: str) -> Optional[str]:
    if not HAS_ZXING or not image_path or not os.path.exists(image_path):
        return None
    try:
        with Image.open(image_path) as img:
            results = zxingcpp.read_barcodes(img)
            if results:
                for r in results:
                    if r.text and r.text.strip():
                        return r.text.strip()

            img_norm = ImageOps.exif_transpose(img) or img
            gray = img_norm.convert("L")
            results = zxingcpp.read_barcodes(gray)
            if results:
                for r in results:
                    if r.text and r.text.strip():
                        return r.text.strip()

            enhancer = ImageEnhance.Contrast(gray)
            enhanced = enhancer.enhance(2.0)
            results = zxingcpp.read_barcodes(enhanced)
            if results:
                for r in results:
                    if r.text and r.text.strip():
                        return r.text.strip()
    except Exception as e:
        logger.warning(f"Barcode detection error: {e}")
    return None


# ---------------------------------------------------------------------------
# 6. GOOGLE SHEETS ASYNC BACKGROUND SYNC
# ---------------------------------------------------------------------------
class SheetsSyncManager:
    def __init__(self, webhook_url: Optional[str] = None):
        self.webhook_url = webhook_url or GOOGLE_SHEET_WEBHOOK_URL
        self._queue: asyncio.Queue[Dict[str, Any]] = asyncio.Queue()
        self._worker_task: Optional[asyncio.Task] = None
        self._session: Optional[aiohttp.ClientSession] = None
        self._is_running = False

    async def start(self):
        if self._is_running:
            return
        self._is_running = True
        self._session = aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10))
        self._worker_task = asyncio.create_task(self._process_queue())

    async def stop(self):
        self._is_running = False
        if self._worker_task:
            self._worker_task.cancel()
            try:
                await self._worker_task
            except asyncio.CancelledError:
                pass
        if self._session:
            await self._session.close()

    def enqueue(self, count_data: Dict[str, Any]):
        if not self.webhook_url:
            return
        self._queue.put_nowait(count_data)

    async def _send(self, data: Dict[str, Any]) -> bool:
        if not self.webhook_url or not self._session:
            return False
        payload = {
            "timestamp": data.get("timestamp", ""),
            "crew": data.get("crew_name", ""),
            "shelf": data.get("shelf", ""),
            "barcode": str(data.get("barcode", "")),
            "name": data.get("item_name", ""),
            "qty": data.get("qty", 1)
        }
        for attempt in range(1, 4):
            try:
                async with self._session.post(self.webhook_url, json=payload) as resp:
                    if resp.status in (200, 201, 302):
                        return True
            except Exception:
                if attempt < 3:
                    await asyncio.sleep(2 ** attempt)
        return False

    async def _process_queue(self):
        while self._is_running:
            try:
                item = await self._queue.get()
                success = await self._send(item)
                if success and "id" in item:
                    await db_mark_synced([item["id"]])
                self._queue.task_done()
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Sync queue error: {e}")
                await asyncio.sleep(1)


sync_manager = SheetsSyncManager()


# ---------------------------------------------------------------------------
# 7. EXCEL EXPORTER (.xlsx)
# ---------------------------------------------------------------------------
def create_excel_report(counts: List[Dict[str, Any]]) -> BytesIO:
    wb = openpyxl.Workbook()
    HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    SUBHEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ZEBRA_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="0F172A")
    BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="1E293B")
    THIN_BORDER = Border(left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"), top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"))
    DOUBLE_BOTTOM = Border(left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"), top=Side(style="thin", color="CBD5E1"), bottom=Side(style="double", color="1E293B"))

    ws = wb.active
    ws.title = "Stock Items"
    ws.views.sheetView[0].showGridLines = True
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"📦 STORE STOCK COUNT REPORT — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["No.", "Date & Time", "Crew Member", "Shelf Location", "Barcode Number", "Item Name / Description", "Quantity", "Sheet Sync"]
    ws.row_dimensions[2].height = 24
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    row_num = 3
    for idx, item in enumerate(counts, 1):
        ws.row_dimensions[row_num].height = 20
        fill = ZEBRA_FILL if idx % 2 == 0 else PatternFill(fill_type=None)
        
        c1 = ws.cell(row=row_num, column=1, value=idx)
        c2 = ws.cell(row=row_num, column=2, value=item.get("timestamp", ""))
        c3 = ws.cell(row=row_num, column=3, value=item.get("crew_name", ""))
        c4 = ws.cell(row=row_num, column=4, value=str(item.get("shelf", "")).upper())
        c4.font = BOLD_FONT

        c5 = ws.cell(row=row_num, column=5, value=str(item.get("barcode", "")))
        c5.number_format = "@"
        c5.font = BOLD_FONT
        c5.alignment = Alignment(horizontal="center", vertical="center")

        c6 = ws.cell(row=row_num, column=6, value=item.get("item_name", "-"))
        c7 = ws.cell(row=row_num, column=7, value=float(item.get("qty", 0)))
        c7.number_format = "#,##0"
        c7.font = BOLD_FONT
        c7.alignment = Alignment(horizontal="right", vertical="center")

        c8 = ws.cell(row=row_num, column=8, value="✅ Synced" if item.get("synced_sheet") == 1 else "⏳ Pending")
        c8.alignment = Alignment(horizontal="center", vertical="center")

        for cell in [c1, c2, c3, c4, c5, c6, c7, c8]:
            if fill.fill_type:
                cell.fill = fill
            cell.border = THIN_BORDER
        row_num += 1

    if counts:
        ws.row_dimensions[row_num].height = 24
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        t_label = ws.cell(row=row_num, column=1, value=f"TOTAL ({len(counts)} SKUs Counted)")
        t_label.font = TOTAL_FONT
        t_label.fill = TOTAL_FILL
        t_label.alignment = Alignment(horizontal="right", vertical="center")

        t_val = ws.cell(row=row_num, column=7, value=f"=SUM(G3:G{row_num-1})")
        t_val.font = TOTAL_FONT
        t_val.fill = TOTAL_FILL
        t_val.number_format = "#,##0"
        t_val.alignment = Alignment(horizontal="right", vertical="center")

        ws.cell(row=row_num, column=8).fill = TOTAL_FILL
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = DOUBLE_BOTTOM
        ws.auto_filter.ref = f"A2:H{row_num-1}"

    # Tab 2: Summary by Shelf
    ws_shelf = wb.create_sheet(title="Summary by Shelf")
    ws_shelf.views.sheetView[0].showGridLines = True
    ws_shelf.merge_cells("A1:C1")
    ws_shelf["A1"].value = "🏢 STOCK BREAKDOWN BY SHELF LOCATION"
    ws_shelf["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws_shelf["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws_shelf["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_shelf.row_dimensions[1].height = 28

    s_headers = ["Shelf Location", "Total SKUs (Items)", "Total Quantity Units"]
    for idx, h in enumerate(s_headers, 1):
        c = ws_shelf.cell(row=2, column=idx, value=h)
        c.fill = SUBHEADER_FILL
        c.font = SUBHEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    shelf_dict = {}
    for item in counts:
        s = str(item.get("shelf", "UNKNOWN")).upper()
        if s not in shelf_dict:
            shelf_dict[s] = {"skus": 0, "qty": 0.0}
        shelf_dict[s]["skus"] += 1
        shelf_dict[s]["qty"] += float(item.get("qty", 0))

    s_row = 3
    for s_name in sorted(shelf_dict.keys()):
        d = shelf_dict[s_name]
        ws_shelf.cell(row=s_row, column=1, value=s_name).font = BOLD_FONT
        ws_shelf.cell(row=s_row, column=2, value=d["skus"]).alignment = Alignment(horizontal="right")
        q = ws_shelf.cell(row=s_row, column=3, value=d["qty"])
        q.alignment = Alignment(horizontal="right")
        q.font = BOLD_FONT
        for col in range(1, 4):
            ws_shelf.cell(row=s_row, column=col).border = THIN_BORDER
        s_row += 1

    # Tab 3: Summary by Crew
    ws_crew = wb.create_sheet(title="Summary by Crew")
    ws_crew.views.sheetView[0].showGridLines = True
    ws_crew.merge_cells("A1:C1")
    ws_crew["A1"].value = "👤 STOCK COUNT ACTIVITY BY CREW MEMBER"
    ws_crew["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws_crew["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws_crew["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_crew.row_dimensions[1].height = 28

    c_headers = ["Crew Member", "Total SKUs Logged", "Total Quantity Logged"]
    for idx, h in enumerate(c_headers, 1):
        c = ws_crew.cell(row=2, column=idx, value=h)
        c.fill = SUBHEADER_FILL
        c.font = SUBHEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    crew_dict = {}
    for item in counts:
        cr = item.get("crew_name") or "Unknown"
        if cr not in crew_dict:
            crew_dict[cr] = {"skus": 0, "qty": 0.0}
        crew_dict[cr]["skus"] += 1
        crew_dict[cr]["qty"] += float(item.get("qty", 0))

    c_row = 3
    for cr_name in sorted(crew_dict.keys(), key=lambda x: crew_dict[x]["qty"], reverse=True):
        d = crew_dict[cr_name]
        ws_crew.cell(row=c_row, column=1, value=cr_name).font = BOLD_FONT
        ws_crew.cell(row=c_row, column=2, value=d["skus"]).alignment = Alignment(horizontal="right")
        q = ws_crew.cell(row=c_row, column=3, value=d["qty"])
        q.alignment = Alignment(horizontal="right")
        q.font = BOLD_FONT
        for col in range(1, 4):
            ws_crew.cell(row=c_row, column=col).border = THIN_BORDER
        c_row += 1

    for worksheet in [ws, ws_shelf, ws_crew]:
        for col in worksheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col if cell.row > 1), default=10)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 8. PARSE 1-SHOT CAPTION
# ---------------------------------------------------------------------------
def parse_quick_caption(caption: str, default_shelf: Optional[str] = None, detected_barcode: Optional[str] = None) -> Optional[Tuple[str, str, str, float]]:
    if not caption or not caption.strip():
        return None
    tokens = caption.strip().split()
    if not tokens:
        return None

    shelf = default_shelf or "UNKNOWN"
    barcode = detected_barcode or "NO_BARCODE"
    item_name = "-"
    qty = 1.0

    if len(tokens) == 1 and tokens[0].replace(".", "", 1).isdigit():
        qty = float(tokens[0])
        return shelf, barcode, item_name, qty

    last_token = tokens[-1]
    if last_token.replace(".", "", 1).isdigit():
        qty = float(last_token)
        tokens = tokens[:-1]

    if not tokens:
        return shelf, barcode, item_name, qty

    if re.match(r"^[A-Za-z0-9\-_]{2,8}$", tokens[0]) and not tokens[0].isdigit():
        shelf = tokens[0].upper()
        tokens = tokens[1:]

    if not tokens:
        return shelf, barcode, item_name, qty

    if tokens[0].isdigit() and len(tokens[0]) >= 6:
        barcode = tokens[0]
        tokens = tokens[1:]

    if tokens:
        item_name = " ".join(tokens)

    return shelf, barcode, item_name, qty


# ---------------------------------------------------------------------------
# 9. TELEGRAM HANDLERS
# ---------------------------------------------------------------------------
def get_user_display_name(update: Update) -> str:
    user = update.effective_user
    if not user:
        return "Unknown"
    full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    if user.username:
        return f"{full_name} (@{user.username})" if full_name else f"@{user.username}"
    return full_name or f"User_{user.id}"


async def save_tg_photo(file_id: str, context: ContextTypes.DEFAULT_TYPE, prefix: str = "img") -> str:
    try:
        tg_file = await context.bot.get_file(file_id)
        filename = f"{prefix}_{uuid.uuid4().hex[:8]}.jpg"
        file_path = str(PHOTOS_DIR / filename)
        await tg_file.download_to_drive(custom_path=file_path)
        return file_path
    except Exception as e:
        logger.error(f"Error saving photo: {e}")
        return ""


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    active_shelf = await db_get_user_active_shelf(user.id) if user else None
    shelf_display = f"`{active_shelf}`" if active_shelf else "_Not set_"

    msg = (
        f"👋 *Welcome to Store Stock Count Bot!*\n\n"
        f"📍 *Your Shelf:* {shelf_display}\n\n"
        f"👉 *How to count an item:*\n"
        f"1️⃣ **Just send a photo of the item** 📸\n"
        f"2️⃣ Bot auto-detects **Name & Barcode** ✨\n"
        f"3️⃣ Type your **Shelf & Quantity**!\n\n"
        f"💡 *Super Fast Pro Tip:*\n"
        f"Send photo with caption: `G101 8850123456789 12`\n"
        f"*(It will save instantly in 1 second!)*"
    )
    await update.message.reply_text(
        msg,
        reply_markup=get_main_reply_keyboard(),
        parse_mode="Markdown"
    )


async def handle_text_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip() if update.message and update.message.text else ""

    if text == "📸 Count New Item":
        await update.message.reply_text("📸 Please send a photo of the product front:", parse_mode="Markdown")
    elif text == "📍 Set Shelf":
        await update.message.reply_text("📍 Please type your **Shelf Code** (e.g. `G101`, `A12`, `B05`):", parse_mode="Markdown")
        context.user_data["waiting_shelf_direct"] = True
    elif text == "📊 Export Excel":
        await cmd_export(update, context)
    elif text == "📈 View Stats":
        await cmd_stats(update, context)
    elif context.user_data.get("waiting_shelf_direct"):
        context.user_data["waiting_shelf_direct"] = False
        shelf = text.upper()
        if update.effective_user:
            await db_set_user_active_shelf(update.effective_user.id, shelf)
        await update.message.reply_text(
            f"✅ *Active shelf set to:* `{shelf}`\n\nAll next items will use `{shelf}` automatically! Send a photo now to count.",
            reply_markup=get_main_reply_keyboard(),
            parse_mode="Markdown"
        )


async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("⏳ *Generating Excel file...*", parse_mode="Markdown")
    try:
        counts = await db_get_all_counts()
        if not counts:
            await msg.edit_text("ℹ️ No items recorded yet.")
            return
        excel_buf = create_excel_report(counts)
        filename = f"Stock_Count_{get_current_timestamp().replace(':', '-').replace(' ', '_')}.xlsx"
        await update.message.reply_document(
            document=excel_buf,
            filename=filename,
            caption=f"📊 *Stock Count Report*\n• Total Items: `{len(counts)}`\n• Time: `{get_current_timestamp()}`",
            reply_markup=get_main_reply_keyboard(),
            parse_mode="Markdown"
        )
        await msg.delete()
    except Exception as e:
        logger.error(f"Export error: {e}")
        await msg.edit_text(f"❌ Error generating Excel: {e}")


async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stats = await db_get_summary_stats()
    text = (
        f"📊 *STOCK COUNT SUMMARY*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 *Total SKUs Counted:* `{stats['total_skus']}`\n"
        f"🔢 *Total Units (Qty):* `{stats['total_qty']:,.0f}`\n"
        f"🏢 *Total Shelves:* `{stats['total_shelves']}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
    )
    if stats["shelf_breakdown"]:
        text += "🏢 *By Shelf:*\n"
        for s in stats["shelf_breakdown"][:6]:
            text += f"• `{s['shelf']}`: {s['sku_count']} SKUs ({s['total_qty']:,.0f} units)\n"
        text += "\n"
    if stats["crew_breakdown"]:
        text += "👤 *By Crew:*\n"
        for c in stats["crew_breakdown"]:
            text += f"• {c['crew_name']}: {c['sku_count']} SKUs ({c['total_qty']:,.0f} units)\n"

    await update.message.reply_text(text, reply_markup=get_main_reply_keyboard(), parse_mode="Markdown")


# ---------------------------------------------------------------------------
# 10. PHOTO FLOW (Auto-Catch Name from Packaging + Barcode + Shelf + QTY)
# ---------------------------------------------------------------------------
async def handle_incoming_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    user = update.effective_user
    crew_name = get_user_display_name(update)

    photo = update.message.photo[-1]
    file_path = await save_tg_photo(photo.file_id, context, prefix="front")
    context.user_data["photo_front"] = file_path

    # Concurrently detect barcode and extract product name from packaging via AI
    detected_barcode, detected_name = await asyncio.gather(
        asyncio.to_thread(detect_barcode_from_image, file_path),
        extract_product_name_from_image(file_path)
    )

    if detected_barcode:
        context.user_data["detected_barcode"] = detected_barcode
    if detected_name:
        context.user_data["item_name"] = detected_name

    active_shelf = await db_get_user_active_shelf(user.id) if user else None
    caption = update.message.caption or ""
    quick_data = parse_quick_caption(caption, default_shelf=active_shelf, detected_barcode=detected_barcode)

    if quick_data:
        shelf, barcode, name, qty = quick_data
        # If quick caption did not specify name, use AI detected name if available
        if name == "-" and detected_name:
            name = detected_name

        record = await db_insert_count(
            user_id=user.id if user else 0,
            crew_name=crew_name,
            shelf=shelf,
            barcode=barcode,
            item_name=name,
            qty=qty,
            photo_front=file_path
        )
        if user and shelf and shelf != "UNKNOWN":
            await db_set_user_active_shelf(user.id, shelf)

        sync_manager.enqueue(record)
        qty_display = int(qty) if qty.is_integer() else qty

        await update.message.reply_text(
            f"⚡ *SAVED! (ID #{record['id']})*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📍 *Shelf:* `{shelf}`\n"
            f"🏷️ *Barcode:* `{barcode}`\n"
            f"📦 *Item:* {name}\n"
            f"🔢 *Quantity:* `{qty_display}`\n"
            f"👤 *Crew:* {crew_name}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👉 *Send next photo to continue!*",
            reply_markup=get_main_reply_keyboard(),
            parse_mode="Markdown"
        )
        return ConversationHandler.END

    name_status = f"\n✨ *Auto-detected Name:* `{detected_name}`" if detected_name else ""

    kb = InlineKeyboardMarkup([[InlineKeyboardButton("⏩ Skip (Barcode is in this photo)", callback_data="skip_barcode_photo")]])
    await update.message.reply_text(
        f"📸 *Photo Received!*{name_status}\n\n"
        f"Send barcode photo (or tap Skip below):",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    return STATE_BARCODE_PHOTO


async def flow_receive_barcode_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message and update.message.photo:
        photo = update.message.photo[-1]
        file_path = await save_tg_photo(photo.file_id, context, prefix="barcode")
        context.user_data["photo_barcode"] = file_path
        detected = detect_barcode_from_image(file_path)
        if detected:
            context.user_data["detected_barcode"] = detected
    return await prompt_shelf_step(update, context)


async def flow_skip_barcode_photo_cb(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    context.user_data["photo_barcode"] = None
    return await prompt_shelf_step(update, context)


async def prompt_shelf_step(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    active_shelf = await db_get_user_active_shelf(user.id) if user else None
    target = update.callback_query.message if update.callback_query else update.message

    if active_shelf:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"✅ Keep Current Shelf: {active_shelf}", callback_data=f"shelf_keep_{active_shelf}")]
        ])
        await target.reply_text(
            f"📍 *Please type the Shelf Code* (e.g. `G101`, `A12`, `B05`):\n\n"
            f"_(Or tap below to keep `{active_shelf}`):_",
            reply_markup=kb,
            parse_mode="Markdown"
        )
    else:
        await target.reply_text(
            "📍 *Please type the Shelf Code* (e.g. `G101`, `A12`, `B05`, etc.):",
            parse_mode="Markdown"
        )
    return STATE_SHELF


async def flow_shelf_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data.startswith("shelf_keep_"):
        shelf = data.replace("shelf_keep_", "").strip().upper()
        context.user_data["shelf"] = shelf
        return await prompt_barcode_step(update, context)

    return STATE_SHELF


async def flow_shelf_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    shelf = update.message.text.strip().upper()
    if not shelf:
        await update.message.reply_text("⚠️ Please type the Shelf Code (e.g. `G101`):")
        return STATE_SHELF
    context.user_data["shelf"] = shelf
    if update.effective_user:
        await db_set_user_active_shelf(update.effective_user.id, shelf)
    return await prompt_barcode_step(update, context)


async def prompt_barcode_step(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    target = update.callback_query.message if update.callback_query else update.message
    detected = context.user_data.get("detected_barcode")
    if detected:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"✅ Confirm Barcode: {detected}", callback_data=f"confirm_barcode_{detected}")],
            [InlineKeyboardButton("✏️ Type Different Barcode", callback_data="type_barcode")]
        ])
        await target.reply_text(f"🏷️ *Barcode Detected:* `{detected}`\nTap to confirm or type manually:", reply_markup=kb, parse_mode="Markdown")
    else:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("⏩ Skip Barcode Number", callback_data="skip_barcode_num")]
        ])
        await target.reply_text("🏷️ Please type the **Barcode numbers** from the label:\n_(Or tap Skip):_", reply_markup=kb, parse_mode="Markdown")
    return STATE_BARCODE


async def flow_barcode_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data.startswith("confirm_barcode_"):
        context.user_data["barcode"] = data.replace("confirm_barcode_", "").strip()
        return await process_item_name_or_skip(update, context)
    elif data == "skip_barcode_num":
        context.user_data["barcode"] = "NO_BARCODE"
        return await process_item_name_or_skip(update, context)
    elif data == "type_barcode":
        await query.message.reply_text("🏷️ Please type the **Barcode number**:", parse_mode="Markdown")
        return STATE_BARCODE
    return STATE_BARCODE


async def flow_barcode_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    barcode = update.message.text.strip()
    if not barcode:
        await update.message.reply_text("⚠️ Barcode cannot be empty.")
        return STATE_BARCODE
    context.user_data["barcode"] = barcode
    return await process_item_name_or_skip(update, context)


async def process_item_name_or_skip(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """If AI already auto-detected the name from packaging, skip asking for name and go straight to Quantity!"""
    detected_name = context.user_data.get("item_name")
    target = update.callback_query.message if update.callback_query else update.message

    if detected_name and detected_name != "-":
        # Name is already caught from packaging! Skip name typing step!
        await target.reply_text(
            f"📦 *Product Name:* `{detected_name}` (Auto-detected ✨)\n\n"
            f"🔢 *Please type the Quantity (QTY):*\n_(e.g. 1, 5, 12, 24)_",
            parse_mode="Markdown"
        )
        return STATE_QTY

    # If AI didn't catch name, give option to type or skip
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("⏩ Skip Product Name", callback_data="skip_item_name")]])
    await target.reply_text("📦 Type the **Product Name** (or tap Skip):", reply_markup=kb, parse_mode="Markdown")
    return STATE_ITEM_NAME


async def flow_item_name_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "skip_item_name":
        context.user_data["item_name"] = "-"
        await query.message.reply_text("🔢 *Please type the Quantity (QTY):*\n_(e.g. 1, 5, 12, 24)_", parse_mode="Markdown")
        return STATE_QTY
    return STATE_ITEM_NAME


async def flow_item_name_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["item_name"] = update.message.text.strip() or "-"
    await update.message.reply_text("🔢 *Please type the Quantity (QTY):*\n_(e.g. 1, 5, 12, 24)_", parse_mode="Markdown")
    return STATE_QTY


async def flow_qty_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip() if update.message and update.message.text else ""
    try:
        qty = float(text)
        if qty <= 0:
            await update.message.reply_text("⚠️ Quantity must be greater than 0. Please type a number (e.g. 12):")
            return STATE_QTY
        context.user_data["qty"] = qty
        return await finalize_and_save_count(update, context)
    except ValueError:
        await update.message.reply_text("⚠️ Please enter a valid number (e.g. `12` or `5`):")
        return STATE_QTY


async def finalize_and_save_count(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    crew_name = get_user_display_name(update)
    shelf = context.user_data.get("shelf", "UNKNOWN")
    barcode = context.user_data.get("barcode", "NO_BARCODE")
    item_name = context.user_data.get("item_name", "-")
    qty = context.user_data.get("qty", 1.0)
    photo_front = context.user_data.get("photo_front")
    photo_barcode = context.user_data.get("photo_barcode")

    record = await db_insert_count(
        user_id=user.id if user else 0,
        crew_name=crew_name,
        shelf=shelf,
        barcode=barcode,
        item_name=item_name,
        qty=qty,
        photo_front=photo_front,
        photo_barcode=photo_barcode
    )

    sync_manager.enqueue(record)
    target = update.callback_query.message if update.callback_query else update.message
    qty_display = int(qty) if qty.is_integer() else qty

    card = (
        f"✅ *STOCK ITEM SAVED! (ID #{record['id']})*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📍 *Shelf:* `{shelf}`\n"
        f"🏷️ *Barcode:* `{barcode}`\n"
        f"📦 *Item:* {item_name}\n"
        f"🔢 *Quantity:* `{qty_display}`\n"
        f"👤 *Crew:* {crew_name}\n"
        f"🕒 *Time:* `{record['timestamp']}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👉 *Send next photo to keep counting on Shelf `{shelf}`!*"
    )
    await target.reply_text(card, reply_markup=get_main_reply_keyboard(), parse_mode="Markdown")
    context.user_data.clear()
    return ConversationHandler.END


async def flow_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text("❌ Cancelled.", reply_markup=get_main_reply_keyboard(), parse_mode="Markdown")
    return ConversationHandler.END


# ---------------------------------------------------------------------------
# 11. LIGHTWEIGHT HTTP HEALTH SERVER
# ---------------------------------------------------------------------------
async def handle_health_check(request):
    if not TELEGRAM_BOT_TOKEN:
        return web.Response(
            text="⚠️ Bot server is online, but TELEGRAM_BOT_TOKEN is missing in Render Environment Variables!",
            content_type="text/plain",
            status=200
        )
    return web.Response(
        text="✅ Store Stock Count Telegram Bot is active and running 24/7!",
        content_type="text/plain",
        status=200
    )


# ---------------------------------------------------------------------------
# 12. MAIN RUNNER
# ---------------------------------------------------------------------------
async def main_async():
    logger.info("Initializing SQLite database...")
    await init_db()
    await sync_manager.start()

    app_web = web.Application()
    app_web.router.add_get("/", handle_health_check)
    app_web.router.add_get("/health", handle_health_check)
    runner = web.AppRunner(app_web)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    logger.info(f"🚀 Web health server listening on port {PORT}")

    if not TELEGRAM_BOT_TOKEN:
        logger.error("❌ ERROR: TELEGRAM_BOT_TOKEN is not set in Environment Variables!")
        while True:
            await asyncio.sleep(3600)

    logger.info("🤖 Starting Telegram Bot polling...")
    tg_app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.PHOTO, handle_incoming_photo),
            CommandHandler("count", lambda u, c: u.message.reply_text("📸 Send a photo of the product front:"))
        ],
        states={
            STATE_BARCODE_PHOTO: [
                MessageHandler(filters.PHOTO, flow_receive_barcode_photo),
                CallbackQueryHandler(flow_skip_barcode_photo_cb, pattern="^skip_barcode_photo$"),
                CommandHandler("cancel", flow_cancel)
            ],
            STATE_SHELF: [
                CallbackQueryHandler(flow_shelf_callback, pattern="^shelf_keep_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, flow_shelf_text),
                CommandHandler("cancel", flow_cancel)
            ],
            STATE_BARCODE: [
                CallbackQueryHandler(flow_barcode_callback, pattern="^(confirm_barcode_|type_barcode|skip_barcode_num)"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, flow_barcode_text),
                CommandHandler("cancel", flow_cancel)
            ],
            STATE_ITEM_NAME: [
                CallbackQueryHandler(flow_item_name_callback, pattern="^skip_item_name$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, flow_item_name_text),
                CommandHandler("cancel", flow_cancel)
            ],
            STATE_QTY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, flow_qty_text),
                CommandHandler("cancel", flow_cancel)
            ]
        },
        fallbacks=[CommandHandler("cancel", flow_cancel)],
        per_user=True,
        per_chat=True
    )

    tg_app.add_handler(conv)
    tg_app.add_handler(CommandHandler("start", cmd_start))
    tg_app.add_handler(CommandHandler("help", cmd_start))
    tg_app.add_handler(CommandHandler("export", cmd_export))
    tg_app.add_handler(CommandHandler("stats", cmd_stats))
    tg_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_buttons))

    await tg_app.initialize()
    await tg_app.start()
    await tg_app.updater.start_polling(drop_pending_updates=True)
    logger.info("🎉 Bot is online and listening for messages!")

    try:
        while True:
            await asyncio.sleep(3600)
    finally:
        await tg_app.updater.stop()
        await tg_app.stop()
        await tg_app.shutdown()
        await sync_manager.stop()
        await runner.cleanup()


def main():
    asyncio.run(main_async())


if __name__ == "__main__":
    main()

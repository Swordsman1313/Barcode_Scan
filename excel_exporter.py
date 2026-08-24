"""
Professional Excel (.xlsx) exporter using openpyxl.
Generates styled, multi-tab workbooks with summary cards, proper barcode formatting (text-safe),
and auto-calculated column widths.
"""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_report(counts: List[Dict[str, Any]], stats: Optional[Dict[str, Any]] = None) -> BytesIO:
    """
    Generates a formatted Excel (.xlsx) report in memory and returns a BytesIO buffer.
    """
    wb = openpyxl.Workbook()

    # Define color palette & styles
    HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    SUBHEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue Accent
    SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ZEBRA_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")   # Light slate
    TOTAL_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")   # Medium slate
    TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="0F172A")

    REGULAR_FONT = Font(name="Calibri", size=11, color="1E293B")
    BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="1E293B")
    
    THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
    THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
    DOUBLE_BOTTOM = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=Side(border_style="double", color="1E293B"))

    # ----------------------------------------------------
    # TAB 1: Detailed Items
    # ----------------------------------------------------
    ws_items = wb.active
    ws_items.title = "Stock Items"
    ws_items.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_items.merge_cells("A1:H1")
    title_cell = ws_items["A1"]
    title_cell.value = f"📦 STORE INVENTORY & STOCK COUNT REPORT — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_items.row_dimensions[1].height = 32

    # Headers
    headers = [
        "No.",
        "Date & Time",
        "Crew Member",
        "Shelf Location",
        "Barcode Number",
        "Item Name / Description",
        "Quantity",
        "Sheet Sync"
    ]

    ws_items.row_dimensions[2].height = 24
    for col_idx, header in enumerate(headers, 1):
        cell = ws_items.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data Rows
    current_row = 3
    for idx, item in enumerate(counts, 1):
        ws_items.row_dimensions[current_row].height = 20
        fill = ZEBRA_FILL if idx % 2 == 0 else PatternFill(fill_type=None)

        c1 = ws_items.cell(row=current_row, column=1, value=idx)
        c1.alignment = Alignment(horizontal="center", vertical="center")

        c2 = ws_items.cell(row=current_row, column=2, value=item.get("timestamp", ""))
        c2.alignment = Alignment(horizontal="center", vertical="center")

        c3 = ws_items.cell(row=current_row, column=3, value=item.get("crew_name", ""))
        c3.alignment = Alignment(horizontal="left", vertical="center")

        c4 = ws_items.cell(row=current_row, column=4, value=str(item.get("shelf", "")).upper())
        c4.alignment = Alignment(horizontal="center", vertical="center")
        c4.font = BOLD_FONT

        # Barcode strictly as text format to prevent scientific notation truncation
        c5 = ws_items.cell(row=current_row, column=5, value=str(item.get("barcode", "")))
        c5.number_format = "@"
        c5.alignment = Alignment(horizontal="center", vertical="center")
        c5.font = BOLD_FONT

        c6 = ws_items.cell(row=current_row, column=6, value=item.get("item_name", "") or "-")
        c6.alignment = Alignment(horizontal="left", vertical="center")

        c7 = ws_items.cell(row=current_row, column=7, value=float(item.get("qty", 0)))
        c7.number_format = "#,##0"
        c7.alignment = Alignment(horizontal="right", vertical="center")
        c7.font = BOLD_FONT

        synced_status = "✅ Synced" if item.get("synced_sheet") == 1 else "⏳ Pending"
        c8 = ws_items.cell(row=current_row, column=8, value=synced_status)
        c8.alignment = Alignment(horizontal="center", vertical="center")

        for cell in [c1, c2, c3, c4, c5, c6, c7, c8]:
            if not cell.font or cell.font == BOLD_FONT:
                pass
            else:
                cell.font = REGULAR_FONT
            if fill.fill_type:
                cell.fill = fill
            cell.border = THIN_BORDER

        current_row += 1

    # Total Summary Row
    if counts:
        ws_items.row_dimensions[current_row].height = 24
        ws_items.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        tot_label = ws_items.cell(row=current_row, column=1, value=f"TOTAL ({len(counts)} SKUs Counted)")
        tot_label.font = TOTAL_FONT
        tot_label.fill = TOTAL_FILL
        tot_label.alignment = Alignment(horizontal="right", vertical="center")
        tot_label.border = DOUBLE_BOTTOM

        tot_val = ws_items.cell(row=current_row, column=7, value=f"=SUM(G3:G{current_row - 1})")
        tot_val.font = TOTAL_FONT
        tot_val.fill = TOTAL_FILL
        tot_val.number_format = "#,##0"
        tot_val.alignment = Alignment(horizontal="right", vertical="center")
        tot_val.border = DOUBLE_BOTTOM

        tot_end = ws_items.cell(row=current_row, column=8, value="")
        tot_end.fill = TOTAL_FILL
        tot_end.border = DOUBLE_BOTTOM

        for col in range(1, 9):
            ws_items.cell(row=current_row, column=col).border = DOUBLE_BOTTOM

        # Auto filter
        ws_items.auto_filter.ref = f"A2:H{current_row - 1}"

    # ----------------------------------------------------
    # TAB 2: Summary by Shelf
    # ----------------------------------------------------
    ws_shelf = wb.create_sheet(title="Summary by Shelf")
    ws_shelf.views.sheetView[0].showGridLines = True
    
    ws_shelf.merge_cells("A1:C1")
    ws_shelf["A1"].value = "🏢 STOCK BREAKDOWN BY SHELF LOCATION"
    ws_shelf["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws_shelf["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws_shelf["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_shelf.row_dimensions[1].height = 28

    shelf_headers = ["Shelf Location", "Total SKUs (Items)", "Total Quantity Units"]
    ws_shelf.row_dimensions[2].height = 22
    for col_idx, h in enumerate(shelf_headers, 1):
        cell = ws_shelf.cell(row=2, column=col_idx, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Aggregate by shelf
    shelf_dict = {}
    for item in counts:
        shelf = str(item.get("shelf", "UNKNOWN")).upper()
        if shelf not in shelf_dict:
            shelf_dict[shelf] = {"skus": 0, "qty": 0.0}
        shelf_dict[shelf]["skus"] += 1
        shelf_dict[shelf]["qty"] += float(item.get("qty", 0))

    s_row = 3
    for shelf_name in sorted(shelf_dict.keys()):
        data = shelf_dict[shelf_name]
        ws_shelf.row_dimensions[s_row].height = 20
        fill = ZEBRA_FILL if s_row % 2 == 0 else PatternFill(fill_type=None)

        c1 = ws_shelf.cell(row=s_row, column=1, value=shelf_name)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.font = BOLD_FONT

        c2 = ws_shelf.cell(row=s_row, column=2, value=data["skus"])
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.number_format = "#,##0"

        c3 = ws_shelf.cell(row=s_row, column=3, value=data["qty"])
        c3.alignment = Alignment(horizontal="right", vertical="center")
        c3.number_format = "#,##0"
        c3.font = BOLD_FONT

        for cell in [c1, c2, c3]:
            if fill.fill_type:
                cell.fill = fill
            cell.border = THIN_BORDER
        s_row += 1

    if shelf_dict:
        ws_shelf.row_dimensions[s_row].height = 22
        t1 = ws_shelf.cell(row=s_row, column=1, value="TOTAL ALL SHELVES")
        t1.font = TOTAL_FONT
        t1.fill = TOTAL_FILL
        t1.alignment = Alignment(horizontal="center", vertical="center")

        t2 = ws_shelf.cell(row=s_row, column=2, value=f"=SUM(B3:B{s_row-1})")
        t2.font = TOTAL_FONT
        t2.fill = TOTAL_FILL
        t2.alignment = Alignment(horizontal="right", vertical="center")
        t2.number_format = "#,##0"

        t3 = ws_shelf.cell(row=s_row, column=3, value=f"=SUM(C3:C{s_row-1})")
        t3.font = TOTAL_FONT
        t3.fill = TOTAL_FILL
        t3.alignment = Alignment(horizontal="right", vertical="center")
        t3.number_format = "#,##0"

        for cell in [t1, t2, t3]:
            cell.border = DOUBLE_BOTTOM

    # ----------------------------------------------------
    # TAB 3: Summary by Crew Member
    # ----------------------------------------------------
    ws_crew = wb.create_sheet(title="Summary by Crew")
    ws_crew.views.sheetView[0].showGridLines = True

    ws_crew.merge_cells("A1:C1")
    ws_crew["A1"].value = "👤 STOCK COUNT ACTIVITY BY CREW MEMBER"
    ws_crew["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws_crew["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws_crew["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_crew.row_dimensions[1].height = 28

    crew_headers = ["Crew Member", "Total SKUs Logged", "Total Quantity Logged"]
    ws_crew.row_dimensions[2].height = 22
    for col_idx, h in enumerate(crew_headers, 1):
        cell = ws_crew.cell(row=2, column=col_idx, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    crew_dict = {}
    for item in counts:
        crew = item.get("crew_name") or "Unknown"
        if crew not in crew_dict:
            crew_dict[crew] = {"skus": 0, "qty": 0.0}
        crew_dict[crew]["skus"] += 1
        crew_dict[crew]["qty"] += float(item.get("qty", 0))

    c_row = 3
    for crew_name in sorted(crew_dict.keys(), key=lambda x: crew_dict[x]["qty"], reverse=True):
        data = crew_dict[crew_name]
        ws_crew.row_dimensions[c_row].height = 20
        fill = ZEBRA_FILL if c_row % 2 == 0 else PatternFill(fill_type=None)

        c1 = ws_crew.cell(row=c_row, column=1, value=crew_name)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.font = BOLD_FONT

        c2 = ws_crew.cell(row=c_row, column=2, value=data["skus"])
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.number_format = "#,##0"

        c3 = ws_crew.cell(row=c_row, column=3, value=data["qty"])
        c3.alignment = Alignment(horizontal="right", vertical="center")
        c3.number_format = "#,##0"
        c3.font = BOLD_FONT

        for cell in [c1, c2, c3]:
            if fill.fill_type:
                cell.fill = fill
            cell.border = THIN_BORDER
        c_row += 1

    if crew_dict:
        ws_crew.row_dimensions[c_row].height = 22
        t1 = ws_crew.cell(row=c_row, column=1, value="TOTAL ALL CREW")
        t1.font = TOTAL_FONT
        t1.fill = TOTAL_FILL
        t1.alignment = Alignment(horizontal="left", vertical="center")

        t2 = ws_crew.cell(row=c_row, column=2, value=f"=SUM(B3:B{c_row-1})")
        t2.font = TOTAL_FONT
        t2.fill = TOTAL_FILL
        t2.alignment = Alignment(horizontal="right", vertical="center")
        t2.number_format = "#,##0"

        t3 = ws_crew.cell(row=c_row, column=3, value=f"=SUM(C3:C{c_row-1})")
        t3.font = TOTAL_FONT
        t3.fill = TOTAL_FILL
        t3.alignment = Alignment(horizontal="right", vertical="center")
        t3.number_format = "#,##0"

        for cell in [t1, t2, t3]:
            cell.border = DOUBLE_BOTTOM

    # ----------------------------------------------------
    # Auto-adjust column widths across all sheets
    # ----------------------------------------------------
    for ws in [ws_items, ws_shelf, ws_crew]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Skip title banner in row 1
                if cell.row == 1:
                    continue
                val_str = str(cell.value or "")
                if val_str.startswith("="):
                    val_str = "123,456"
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
